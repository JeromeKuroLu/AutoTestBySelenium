import os
from openpyxl import Workbook, load_workbook


class Xlsx_Util:
    def __init__(self, dir):
        super().__init__()
        self.file_dir = dir

    def write(self, title_array, data_array):
        if os.path.exists(self.file_dir):
            wb = load_workbook(self.file_dir)
            sheet = wb.active
            sheet.append(data_array)
        else:
            wb = Workbook()
            sheet = wb.active
            sheet.append(title_array)
            sheet.append(data_array)
        wb.save(self.file_dir)